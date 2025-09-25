import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import timedelta, date
from pathlib import Path
import logging
from typing import List, Optional, Dict

logger = logging.getLogger(__name__)


class ExcelTemplateManager:
    """Create and manage Excel tracking templates"""

    def __init__(self, excel_directory: str = "excel_files"):
        self.excel_directory = Path(excel_directory)
        self.excel_directory.mkdir(exist_ok=True)

    def create_excel_template(self, session_id: int, data: Dict,
                              username: str) -> Optional[str]:
        """Create Excel template with date headers and reply rows"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reply Tracking"

            # Generate date range
            start_date = data['start_date']
            end_date = data['end_date']
            dates = self._generate_date_range(start_date, end_date)

            # Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1DA1F2",
                                      end_color="1DA1F2",
                                      fill_type="solid")
            center_alignment = Alignment(horizontal="center",
                                         vertical="center")

            # Create Reply # column (first column)
            ws.cell(row=1, column=1).value = "Reply #"
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).alignment = center_alignment
            ws.column_dimensions['A'].width = 10

            # Create date headers (starting from column 2)
            for col_idx, date_obj in enumerate(dates, start=2):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = date_obj.strftime('%Y-%m-%d')
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

            # Create reply number rows
            for reply_num in range(1, data['target_replies'] + 1):
                cell = ws.cell(row=reply_num + 1, column=1)
                cell.value = reply_num
                cell.alignment = center_alignment
                cell.font = Font(bold=True)
                for col_idx in range(2, len(dates) + 2):
                    cell = ws.cell(row=reply_num + 1, column=col_idx)
                    cell.value = ""
                    cell.alignment = center_alignment

            # Save file
            safe_username = "".join(
                c for c in username
                if c.isalnum() or c in (' ', '-', '_')).strip()
            filename = f"tracking_{session_id}_{safe_username.replace(' ', '_')}.xlsx"
            filepath = self.excel_directory / filename

            wb.save(str(filepath))
            logger.info(f"Excel template created: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error creating Excel template: {e}", exc_info=True)
            return None

    def update_excel_file(self, excel_path: str, session_id: int,
                          date_obj: date, urls: List[str], target_replies: int,
                          start_date: date, x_username: str) -> bool:
        """Update Excel file with new reply links"""
        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excel file not found: {excel_path}")
                return False

            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            date_str = date_obj.strftime('%Y-%m-%d')
            date_col = None

            # Search for date in header row (row 1, starting from column 2)
            for col in range(2, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if isinstance(cell_value, str) and cell_value == date_str:
                    date_col = col
                    break

            if not date_col:
                logger.error(f"Date column {date_str} not found in Excel")
                return False

            # Clear existing content in this column (except header)
            for row in range(2, target_replies + 2):
                cell = ws.cell(row=row, column=date_col)
                cell.value = ""
                cell.hyperlink = None

            # Add all replies as hyperlinks
            for idx, url in enumerate(urls):
                if idx < target_replies:
                    row_num = idx + 2
                    cell = ws.cell(row=row_num, column=date_col)
                    cell.value = str(idx + 1)
                    cell.hyperlink = url
                    cell.font = Font(color="0000FF", underline="single")

            wb.save(excel_path)
            logger.info(f"Excel updated: {len(urls)} replies for {date_str}")
            return True

        except Exception as e:
            logger.error(f"Error updating Excel file: {e}", exc_info=True)
            return False

    def _generate_date_range(self, start_date: date,
                             end_date: date) -> List[date]:
        """Generate list of dates between start and end date"""
        dates = []
        current_date = start_date
        while current_date <= end_date:
            dates.append(current_date)
            current_date = current_date + timedelta(days=1)
        return dates
