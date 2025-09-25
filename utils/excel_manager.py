import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, date, timedelta
import logging
from typing import Optional, List, Dict

logger = logging.getLogger(__name__)

class CombinedExcelReportGenerator:
    """Generate combined Excel reports with multiple sheets"""

    def __init__(self, db_manager):
        self.db = db_manager

    async def generate_combined_report(self) -> Optional[str]:
        """Generate a single Excel file with multiple sheets for all users"""
        try:
            async with self.db.get_db() as db:
                async with db.execute('''
                    SELECT u.id, u.discord_id, u.username, u.x_username, 
                           ts.id as session_id, ts.target_replies, ts.start_date, ts.end_date,
                           ts.excel_path, COUNT(r.id) as total_replies
                    FROM users u
                    JOIN tracking_sessions ts ON u.id = ts.user_id
                    LEFT JOIN replies r ON ts.id = r.session_id AND r.is_valid = 1
                    WHERE ts.status = 'active'
                    GROUP BY u.id, ts.id
                    ORDER BY u.username
                ''') as cursor:
                    users_data = await cursor.fetchall()

            if not users_data:
                return None

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            summary_sheet = wb.create_sheet("📊 Summary")
            await self._create_summary_sheet(summary_sheet, users_data)

            for user_data in users_data:
                await self._create_user_sheet(wb, user_data)

            analytics_sheet = wb.create_sheet("📈 Analytics")
            await self._create_analytics_sheet(analytics_sheet, users_data)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"combined_reports_{timestamp}.xlsx"
            filepath = os.path.join("excel_files", filename)
            os.makedirs("excel_files", exist_ok=True)
            wb.save(filepath)
            logger.info(f"Generated combined report: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error generating combined report: {e}", exc_info=True)
            return None

    async def _create_summary_sheet(self, sheet, users_data: List[Dict]):
        # Title
        sheet.merge_cells('A1:H1')
        title_cell = sheet['A1']
        title_cell.value = f"Reply Tracking Summary - Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1DA1F2", end_color="1DA1F2", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = ["Username", "X Username", "Target/Day", "Period", "Total Replies", "Avg/Day", "Completion %", "Status"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, user_data in enumerate(users_data, 4):
            start_date = datetime.strptime(user_data['start_date'], '%Y-%m-%d').date()
            end_date = datetime.strptime(user_data['end_date'], '%Y-%m-%d').date()
            total_days = (end_date - start_date).days + 1
            today = date.today()

            if today < start_date:
                days_elapsed = 0
                status = "Not Started"
            elif today > end_date:
                days_elapsed = total_days
                status = "Completed"
            else:
                days_elapsed = (today - start_date).days + 1
                status = "Active"

            expected_replies = days_elapsed * user_data['target_replies']
            completion_pct = (user_data['total_replies'] / expected_replies * 100) if expected_replies > 0 else 0
            avg_per_day = user_data['total_replies'] / days_elapsed if days_elapsed > 0 else 0

            sheet.cell(row=row_idx, column=1).value = user_data['username']
            sheet.cell(row=row_idx, column=2).value = f"@{user_data['x_username']}"
            sheet.cell(row=row_idx, column=3).value = user_data['target_replies']
            sheet.cell(row=row_idx, column=4).value = f"{start_date} to {end_date}"
            sheet.cell(row=row_idx, column=5).value = user_data['total_replies']
            sheet.cell(row=row_idx, column=6).value = round(avg_per_day, 1)
            sheet.cell(row=row_idx, column=7).value = f"{completion_pct:.1f}%"
            sheet.cell(row=row_idx, column=8).value = status

            completion_cell = sheet.cell(row=row_idx, column=7)
            if completion_pct >= 100:
                completion_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif completion_pct >= 80:
                completion_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                completion_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    async def _create_user_sheet(self, workbook, user_data: Dict):
        try:
            safe_name = "".join(c for c in user_data['username'] if c.isalnum() or c in (' ', '-', '_'))[:25]
            sheet_name = f"{safe_name}"
            sheet = workbook.create_sheet(sheet_name)

            async with self.db.get_db() as db:
                async with db.execute('''
                    SELECT date, COUNT(*) as replies_count,
                           GROUP_CONCAT(url, '||') as urls
                    FROM replies 
                    WHERE session_id = ? AND is_valid = 1
                    GROUP BY date
                    ORDER BY date
                ''', (user_data['session_id'],)) as cursor:
                    reply_data = await cursor.fetchall()

            start_date = datetime.strptime(user_data['start_date'], '%Y-%m-%d').date()
            end_date = datetime.strptime(user_data['end_date'], '%Y-%m-%d').date()
            dates = []
            current_date = start_date
            while current_date <= end_date:
                dates.append(current_date)
                current_date = current_date + timedelta(days=1)

            sheet.merge_cells('A1:C1')
            info_cell = sheet['A1']
            info_cell.value = f"{user_data['username']} (@{user_data['x_username']}) - Target: {user_data['target_replies']}/day"
            info_cell.font = Font(size=12, bold=True)
            info_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            sheet.cell(row=3, column=1, value="Reply #").font = Font(bold=True)
            for col_idx, date_obj in enumerate(dates[:30], 2):
                cell = sheet.cell(row=3, column=col_idx)
                cell.value = date_obj.strftime('%m-%d')
                cell.font = Font(bold=True)
                sheet.column_dimensions[get_column_letter(col_idx)].width = 8

            reply_dict = {row['date']: row for row in reply_data}

            for reply_num in range(1, min(user_data['target_replies'] + 1, 51)):
                sheet.cell(row=reply_num + 3, column=1, value=reply_num)

                for col_idx, date_obj in enumerate(dates[:30], 2):
                    date_str = date_obj.strftime('%Y-%m-%d')
                    cell = sheet.cell(row=reply_num + 3, column=col_idx)
                    if date_str in reply_dict:
                        urls = reply_dict[date_str]['urls'].split('||') if reply_dict[date_str]['urls'] else []
                        if len(urls) >= reply_num:
                            cell.value = "✓"
                            cell.hyperlink = urls[reply_num - 1]
                            cell.font = Font(color="0000FF")
                            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        except Exception as e:
            logger.error(f"Error creating user sheet for {user_data['username']}: {e}")

    async def _create_analytics_sheet(self, sheet, users_data: List[Dict]):
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = "Analytics & Insights"
        title_cell.font = Font(size=14, bold=True)
        title_cell.fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")

        row = 3
        total_users = len(users_data)
        total_replies = sum(user['total_replies'] for user in users_data)
        avg_target = sum(user['target_replies'] for user in users_data) / total_users if total_users > 0 else 0

        stats = [
            ("Total Active Users", total_users),
            ("Total Replies Submitted", total_replies),
            ("Average Daily Target", f"{avg_target:.1f}"),
            ("Average Replies per User", f"{total_replies / total_users:.1f}" if total_users > 0 else "0"),
        ]

        sheet.cell(row=row, column=1, value="Overall Statistics").font = Font(bold=True, size=12)
        row += 1

        for stat_name, stat_value in stats:
            sheet.cell(row=row, column=1, value=stat_name)
            sheet.cell(row=row, column=2, value=stat_value)
            row += 1

        row += 2

        sorted_users = sorted(users_data, key=lambda x: x['total_replies'], reverse=True)
        sheet.cell(row=row, column=1, value="Top Performers").font = Font(bold=True, size=12)
        row += 1

        for i, user in enumerate(sorted_users[:5], 1):
            sheet.cell(row=row, column=1, value=f"{i}. {user['username']}")
            sheet.cell(row=row, column=2, value=f"{user['total_replies']} replies")
            row += 1