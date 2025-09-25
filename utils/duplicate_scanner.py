import re
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
from collections import defaultdict
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill
import os
from datetime import datetime

logger = logging.getLogger(__name__)


@dataclass
class DuplicateInfo:
    tweet_id: str
    url: str
    dates: List[str]
    reply_numbers: List[int]
    user_name: str


class AdvancedDuplicateScanner:
    """Advanced duplicate detection with cross-user analysis"""

    def __init__(self, db_manager):
        self.db = db_manager
        self.tweet_id_pattern = re.compile(r'/status/(\d+)')

    def extract_tweet_id(self, url: str) -> Optional[str]:
        """Extract tweet ID from URL"""
        match = self.tweet_id_pattern.search(url)
        return match.group(1) if match else None

    async def scan_users_for_duplicates(self,
                                        user_ids: List[int]) -> Dict[str, Any]:
        """Scan multiple users for duplicate submissions"""
        try:
            results = {
                'users_scanned': [],
                'duplicates_found': [],
                'cross_user_duplicates': [],
                'summary': {}
            }

            for user_id in user_ids:
                user_result = await self._scan_single_user(user_id)
                if user_result:
                    results['users_scanned'].append(user_result)

            cross_duplicates = await self._detect_cross_user_duplicates(
                user_ids)
            results['cross_user_duplicates'] = cross_duplicates

            total_duplicates = sum(
                len(user['internal_duplicates'])
                for user in results['users_scanned'])
            results['summary'] = {
                'total_users_scanned':
                len(results['users_scanned']),
                'total_internal_duplicates':
                total_duplicates,
                'cross_user_duplicates':
                len(cross_duplicates),
                'users_with_issues':
                len([
                    u for u in results['users_scanned']
                    if u['internal_duplicates']
                ]) + len(cross_duplicates)
            }

            return results

        except Exception as e:
            logger.error(f"Error scanning users for duplicates: {e}",
                         exc_info=True)
            return {'error': str(e)}

    async def _scan_single_user(self, discord_id: int) -> Optional[Dict]:
        """Scan individual user for internal duplicates"""
        try:
            async with self.db.get_db() as db:
                async with db.execute(
                        '''
                    SELECT u.id, u.username, u.x_username, ts.id as session_id, ts.target_replies
                    FROM users u
                    JOIN tracking_sessions ts ON u.id = ts.user_id
                    WHERE u.discord_id = ? AND ts.status = 'active'
                    ORDER BY ts.created_at DESC
                    LIMIT 1
                ''', (discord_id, )) as cursor:
                    user_data = await cursor.fetchone()

                if not user_data:
                    return None

                # Get all replies for this user and include username in the result rows
                async with db.execute(
                        '''
                    SELECT date, url, reply_number, created_at
                    FROM replies
                    WHERE session_id = ? AND is_valid = 1
                    ORDER BY date, reply_number
                ''', (user_data['session_id'], )) as cursor:
                    replies = await cursor.fetchall()

            # Pass username explicitly to fix the missing user_name in DuplicateInfo
            duplicates = self._find_internal_duplicates(
                replies, user_data['username'])

            return {
                'discord_id': discord_id,
                'username': user_data['username'],
                'x_username': user_data['x_username'],
                'total_replies': len(replies),
                'internal_duplicates': duplicates,
                'duplicate_count': len(duplicates)
            }

        except Exception as e:
            logger.error(f"Error scanning user {discord_id}: {e}")
            return {
                'discord_id': discord_id,
                'username': 'Unknown',
                'error': str(e),
                'total_replies': 0,
                'internal_duplicates': [],
                'duplicate_count': 0
            }

    def _find_internal_duplicates(self, replies,
                                  username: str) -> List[DuplicateInfo]:
        """Find duplicate URLs within a single user's submissions"""
        url_tracker = defaultdict(list)
        tweet_id_tracker = defaultdict(list)

        for reply in replies:
            url = reply['url']
            tweet_id = self.extract_tweet_id(url)

            url_tracker[url].append(reply)
            if tweet_id:
                tweet_id_tracker[tweet_id].append(reply)

        duplicates = []
        for url, reply_list in url_tracker.items():
            if len(reply_list) > 1:
                duplicates.append(
                    DuplicateInfo(
                        tweet_id=self.extract_tweet_id(url) or 'unknown',
                        url=url,
                        dates=[reply['date'] for reply in reply_list],
                        reply_numbers=[
                            reply['reply_number'] for reply in reply_list
                        ],
                        user_name=username))

        return duplicates

    async def _detect_cross_user_duplicates(self,
                                            user_ids: List[int]) -> List[Dict]:
        """Detect if users are submitting the same tweets"""
        try:
            cross_duplicates = []

            async with self.db.get_db() as db:
                placeholders = ','.join('?' * len(user_ids))
                async with db.execute(
                        f'''
                    SELECT u.username, u.x_username, r.url, r.date, r.reply_number
                    FROM replies r
                    JOIN tracking_sessions ts ON r.session_id = ts.id
                    JOIN users u ON ts.user_id = u.id
                    WHERE u.discord_id IN ({placeholders}) AND r.is_valid = 1
                    ORDER BY r.url
                ''', user_ids) as cursor:
                    all_replies = await cursor.fetchall()

            tweet_groups = defaultdict(list)
            for reply in all_replies:
                tweet_id = self.extract_tweet_id(reply['url'])
                if tweet_id:
                    tweet_groups[tweet_id].append(reply)

            for tweet_id, reply_list in tweet_groups.items():
                users_with_tweet = defaultdict(list)
                for reply in reply_list:
                    user_key = f"{reply['username']} (@{reply['x_username']})"
                    users_with_tweet[user_key].append(reply)

                if len(users_with_tweet) > 1:
                    cross_duplicates.append({
                        'tweet_id':
                        tweet_id,
                        'url':
                        reply_list[0]['url'],
                        'users_affected':
                        dict(users_with_tweet),
                        'total_submissions':
                        len(reply_list)
                    })

            return cross_duplicates

        except Exception as e:
            logger.error(f"Error detecting cross-user duplicates: {e}")
            return []

    async def generate_duplicate_report_file(
            self, scan_results: Dict[str, Any]) -> Optional[str]:
        """Generate detailed Excel report of duplicate findings"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            summary_sheet = wb.create_sheet("Summary")
            self._create_summary_sheet(summary_sheet, scan_results)

            if any(user['internal_duplicates']
                   for user in scan_results['users_scanned']):
                internal_sheet = wb.create_sheet("Internal Duplicates")
                self._create_internal_duplicates_sheet(internal_sheet,
                                                       scan_results)

            if scan_results['cross_user_duplicates']:
                cross_sheet = wb.create_sheet("Cross-User Duplicates")
                self._create_cross_duplicates_sheet(cross_sheet, scan_results)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"duplicate_scan_report_{timestamp}.xlsx"
            filepath = os.path.join("excel_files", filename)

            os.makedirs("excel_files", exist_ok=True)
            wb.save(filepath)

            return filepath

        except Exception as e:
            logger.error(f"Error generating duplicate report file: {e}")
            return None

    def _create_summary_sheet(self, sheet, results):
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = f"Duplicate Scan Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="E74C3C",
                                      end_color="E74C3C",
                                      fill_type="solid")

        summary = results['summary']
        row = 3

        stats = [("Users Scanned", summary['total_users_scanned']),
                 ("Internal Duplicates Found",
                  summary['total_internal_duplicates']),
                 ("Cross-User Duplicates", summary['cross_user_duplicates']),
                 ("Users with Issues", summary['users_with_issues'])]

        for stat_name, stat_value in stats:
            sheet.cell(row=row, column=1,
                       value=stat_name).font = Font(bold=True)
            sheet.cell(row=row, column=2, value=stat_value)
            row += 1

    def _create_internal_duplicates_sheet(self, sheet, results):
        headers = [
            "User", "Tweet ID", "URL", "Dates", "Reply Numbers", "Occurrences"
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F39C12",
                                    end_color="F39C12",
                                    fill_type="solid")

        row = 2
        for user in results['users_scanned']:
            for duplicate in user['internal_duplicates']:
                sheet.cell(row=row, column=1, value=user['username'])
                sheet.cell(row=row, column=2, value=duplicate.tweet_id)
                sheet.cell(row=row, column=3, value=duplicate.url)
                sheet.cell(row=row, column=4, value=', '.join(duplicate.dates))
                sheet.cell(row=row,
                           column=5,
                           value=', '.join(map(str, duplicate.reply_numbers)))
                sheet.cell(row=row, column=6, value=len(duplicate.dates))
                row += 1

    def _create_cross_duplicates_sheet(self, sheet, results):
        headers = ["Tweet ID", "URL", "Users Affected", "Total Submissions"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="9B59B6",
                                    end_color="9B59B6",
                                    fill_type="solid")

        row = 2
        for duplicate in results['cross_user_duplicates']:
            sheet.cell(row=row, column=1, value=duplicate['tweet_id'])
            sheet.cell(row=row, column=2, value=duplicate['url'])

            users_list = []
            for user, replies in duplicate['users_affected'].items():
                users_list.append(f"{user} ({len(replies)} times)")

            sheet.cell(row=row, column=3, value='; '.join(users_list))
            sheet.cell(row=row, column=4, value=duplicate['total_submissions'])
            row += 1
